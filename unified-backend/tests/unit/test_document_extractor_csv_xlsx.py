"""Unit tests for CSV and XLSX extraction in the document_extractor module."""

import csv
import os
import tempfile

import pytest

from app.services.document_extractor import extract_csv, extract_xlsx, MAX_DATA_ROWS


# --- CSV Tests ---


@pytest.fixture
def simple_csv(tmp_path):
    """Create a simple CSV with headers and a few rows."""
    csv_path = tmp_path / "simple.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Age", "City"])
        writer.writerow(["Alice", "30", "New York"])
        writer.writerow(["Bob", "25", "London"])
        writer.writerow(["Charlie", "35", "Tokyo"])
    return str(csv_path)


@pytest.fixture
def empty_csv(tmp_path):
    """Create an empty CSV file."""
    csv_path = tmp_path / "empty.csv"
    csv_path.write_text("")
    return str(csv_path)


@pytest.fixture
def large_csv(tmp_path):
    """Create a CSV with more than 5000 rows."""
    csv_path = tmp_path / "large.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Value"])
        for i in range(6000):
            writer.writerow([str(i), f"value_{i}"])
    return str(csv_path)


@pytest.fixture
def csv_with_special_chars(tmp_path):
    """Create a CSV with special characters including pipes and commas."""
    csv_path = tmp_path / "special.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Description"])
        writer.writerow(["Test", "Has a | pipe character"])
        writer.writerow(["Quoted", 'She said "hello"'])
    return str(csv_path)


@pytest.mark.asyncio
async def test_extract_csv_basic(simple_csv):
    """Test basic CSV extraction with headers and rows."""
    result = await extract_csv(simple_csv)
    lines = result.split("\n")

    # Header line
    assert lines[0] == "Name | Age | City"
    # Separator
    assert lines[1] == "---"
    # Data rows
    assert lines[2] == "Alice | 30 | New York"
    assert lines[3] == "Bob | 25 | London"
    assert lines[4] == "Charlie | 35 | Tokyo"
    assert len(lines) == 5  # header + separator + 3 data rows


@pytest.mark.asyncio
async def test_extract_csv_empty(empty_csv):
    """Test empty CSV raises ValueError."""
    with pytest.raises(ValueError, match="empty"):
        await extract_csv(empty_csv)


@pytest.mark.asyncio
async def test_extract_csv_row_limit(large_csv):
    """Test CSV extraction limits to 5000 data rows."""
    result = await extract_csv(large_csv)
    lines = result.split("\n")

    # header + separator + MAX_DATA_ROWS data rows
    assert len(lines) == 2 + MAX_DATA_ROWS
    # First data row should be present
    assert "0 | value_0" in lines[2]
    # Last allowed row (index 4999) should be present
    assert "4999 | value_4999" in lines[-1]


@pytest.mark.asyncio
async def test_extract_csv_preserves_headers_when_truncated(large_csv):
    """Test that headers are always preserved even when rows are truncated."""
    result = await extract_csv(large_csv)
    lines = result.split("\n")
    assert lines[0] == "ID | Value"


@pytest.mark.asyncio
async def test_extract_csv_special_characters(csv_with_special_chars):
    """Test CSV with special characters in values."""
    result = await extract_csv(csv_with_special_chars)
    assert "Name | Description" in result
    assert "Has a | pipe character" in result


@pytest.mark.asyncio
async def test_extract_csv_nonexistent_file():
    """Test non-existent file raises ValueError."""
    with pytest.raises((ValueError, FileNotFoundError)):
        await extract_csv("/nonexistent/path/to/file.csv")


@pytest.mark.asyncio
async def test_extract_csv_header_only(tmp_path):
    """Test CSV with only a header row (no data)."""
    csv_path = tmp_path / "header_only.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Col1", "Col2", "Col3"])
    result = await extract_csv(str(csv_path))
    lines = result.split("\n")
    assert lines[0] == "Col1 | Col2 | Col3"
    assert lines[1] == "---"
    assert len(lines) == 2  # header + separator only


# --- XLSX Tests ---


@pytest.fixture
def simple_xlsx(tmp_path):
    """Create a simple XLSX with headers and a few rows."""
    import openpyxl

    xlsx_path = tmp_path / "simple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "London"])
    ws.append(["Charlie", 35, "Tokyo"])
    wb.save(xlsx_path)
    return str(xlsx_path)


@pytest.fixture
def empty_xlsx(tmp_path):
    """Create an XLSX with no data."""
    import openpyxl

    xlsx_path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    # The default sheet has no data but iter_rows won't yield anything
    # unless there's at least one cell. Let's create a truly empty workbook.
    ws = wb.active
    wb.save(xlsx_path)
    return str(xlsx_path)


@pytest.fixture
def large_xlsx(tmp_path):
    """Create an XLSX with more than 5000 rows."""
    import openpyxl

    xlsx_path = tmp_path / "large.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Value"])
    for i in range(6000):
        ws.append([i, f"value_{i}"])
    wb.save(xlsx_path)
    return str(xlsx_path)


@pytest.fixture
def corrupted_xlsx(tmp_path):
    """Create a corrupted XLSX file (just random bytes)."""
    xlsx_path = tmp_path / "corrupted.xlsx"
    xlsx_path.write_bytes(b"This is not a valid XLSX file at all")
    return str(xlsx_path)


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """Create an XLSX with multiple sheets to verify only first sheet is read."""
    import openpyxl

    xlsx_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["First", "Sheet"])
    ws1.append(["data1", "data2"])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Second", "Sheet"])
    ws2.append(["other1", "other2"])

    wb.save(xlsx_path)
    return str(xlsx_path)


@pytest.mark.asyncio
async def test_extract_xlsx_basic(simple_xlsx):
    """Test basic XLSX extraction with headers and rows."""
    result = await extract_xlsx(simple_xlsx)
    lines = result.split("\n")

    # Header line
    assert lines[0] == "Name | Age | City"
    # Separator
    assert lines[1] == "---"
    # Data rows (numbers are stringified)
    assert lines[2] == "Alice | 30 | New York"
    assert lines[3] == "Bob | 25 | London"
    assert lines[4] == "Charlie | 35 | Tokyo"
    assert len(lines) == 5


@pytest.mark.asyncio
async def test_extract_xlsx_empty(empty_xlsx):
    """Test empty XLSX raises ValueError."""
    with pytest.raises(ValueError, match="empty"):
        await extract_xlsx(empty_xlsx)


@pytest.mark.asyncio
async def test_extract_xlsx_row_limit(large_xlsx):
    """Test XLSX extraction limits to 5000 data rows."""
    result = await extract_xlsx(large_xlsx)
    lines = result.split("\n")

    # header + separator + MAX_DATA_ROWS data rows
    assert len(lines) == 2 + MAX_DATA_ROWS


@pytest.mark.asyncio
async def test_extract_xlsx_preserves_headers_when_truncated(large_xlsx):
    """Test that headers are always preserved even when rows are truncated."""
    result = await extract_xlsx(large_xlsx)
    lines = result.split("\n")
    assert lines[0] == "ID | Value"


@pytest.mark.asyncio
async def test_extract_xlsx_corrupted(corrupted_xlsx):
    """Test corrupted XLSX raises ValueError."""
    with pytest.raises(ValueError, match="Failed to parse XLSX"):
        await extract_xlsx(corrupted_xlsx)


@pytest.mark.asyncio
async def test_extract_xlsx_nonexistent_file():
    """Test non-existent file raises ValueError."""
    with pytest.raises((ValueError, FileNotFoundError)):
        await extract_xlsx("/nonexistent/path/to/file.xlsx")


@pytest.mark.asyncio
async def test_extract_xlsx_first_sheet_only(multi_sheet_xlsx):
    """Test that only the first (active) sheet is read."""
    result = await extract_xlsx(multi_sheet_xlsx)
    assert "First | Sheet" in result
    assert "data1 | data2" in result
    # Should NOT contain second sheet data
    assert "Second" not in result
    assert "other1" not in result


@pytest.mark.asyncio
async def test_extract_xlsx_none_cells(tmp_path):
    """Test that None cells are rendered as empty strings."""
    import openpyxl

    xlsx_path = tmp_path / "none_cells.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append(["val1", None, "val3"])
    wb.save(xlsx_path)

    result = await extract_xlsx(str(xlsx_path))
    lines = result.split("\n")
    assert lines[2] == "val1 |  | val3"
